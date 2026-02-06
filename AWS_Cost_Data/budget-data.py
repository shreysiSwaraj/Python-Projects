import openpyxl

fileA = openpyxl.load_workbook("fileA.xlsx")
fileB = openpyxl.load_workbook("fileB.xlsx")    #downloaded one

sheet_a = fileA["Sheet1"]
sheet_b = fileB["Sheet1"]

service_list = {}

#Improvised logic
#iter_rows(values_only=True) pulls raw values in bulk
#.cell() accesses cells one by one (slower)
#values_only=True gives raw data, not cell object

for service_row in sheet_b.iter_rows(min_row=2, values_only=True):
    service_name = service_row[0]
    cost = service_row[1]
    service_list[service_name] = cost

#before logic of above for loop
# for service_row in range(2, sheet_b.max_row + 1):
#     service_name = sheet_b.cell(service_row, 1).value
#     cost = sheet_b.cell(service_row, 2).value
#     service_list[service_name] = cost


#very important:Improvised logic
#cant use values_only= true here because service_row is a tuple which is immutable
#when using service_row[0].value,I am NOT modifying the tuple, I am modifying the Cell objects inside the tuple.
for service_row in sheet_a.iter_rows(min_row=2):
    service_name = service_row[0].value
    if service_name in service_list:
        service_row[1].value = service_list[service_name]

#previous logic of above for loop
# for service_row in range(2, sheet_a.max_row + 1):
#     service_name = sheet_a.cell(service_row, 1).value
#     if service_name in service_list:
#         sheet_a.cell(service_row, 2).value = service_list[service_name]


fileA.save("fileA.xlsx")