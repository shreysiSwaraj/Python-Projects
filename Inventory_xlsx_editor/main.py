import openpyxl

inv_file = openpyxl.load_workbook("inventory_details.xlsx")
product_list = inv_file["Sheet1"]
products_per_supplier = {}
inv_per_supplier = {}
inv_under10 = {}


for product_row in range(2, product_list.max_row + 1):

    supplier_name = product_list.cell(product_row, 4).value
    inventory_value = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    inventory_price = product_list.cell(product_row, 5)
    #print(supplier_name)


#calculating products per supplier
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] = products_per_supplier[supplier_name] + 1
    else:
        #print("Adding a new supplier")
        products_per_supplier[supplier_name] = 1



#calculating the inventory value per supplier
    if supplier_name in inv_per_supplier:
        inv_per_supplier[supplier_name] = inventory_value * price + inv_per_supplier[supplier_name]
    else:
        #print(supplier_name)
        inv_per_supplier[supplier_name] = inventory_value * price

#createing a new column of inventory price
    inventory_price.value = inventory_value * price



#finding all the products with inventory less than 10

    product = product_list.cell(product_row, 1).value

    if inventory_value < 200:
        inv_under10[product] = inventory_value



print(products_per_supplier)
print(inv_per_supplier)
print(inv_under10)
product_list.cell(1, 5).value = "Inventory"

inv_file.save("inventory_updated.xlsx")